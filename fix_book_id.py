import pandas as pd
import openpyxl
import re
from openpyxl import load_workbook

# Загружаем файл Excel
file_path = r'd:\svet\nebsvet 14.11.2025.xlsx'

# Загружаем workbook для работы со всеми листами
wb = load_workbook(file_path)

# Получаем список всех листов
sheet_names = wb.sheetnames
print(f"Найдено листов: {len(sheet_names)}")
print(f"Имена листов: {sheet_names}")

# Словарь для хранения всех DataFrame
dataframes = {}

# Обрабатываем каждый лист
for sheet_name in sheet_names:
    print(f"\nОбработка листа: {sheet_name}")
    
    # Читаем лист в DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Выводим информацию о структуре
    print(f"Колонки: {list(df.columns)}")
    print(f"Количество строк: {len(df)}")
    
    # Ищем колонку с автором (может быть "Автор" или похожее)
    author_col = None
    for col in df.columns:
        if 'автор' in str(col).lower() or 'author' in str(col).lower():
            author_col = col
            break
    
    if author_col is None:
        print(f"  Предупреждение: не найдена колонка с автором на листе {sheet_name}")
        continue
    
    print(f"  Найдена колонка автора: {author_col}")
    
    # Ищем или создаем колонку book_id
    book_id_col = None
    for col in df.columns:
        if 'book_id' in str(col).lower() or 'bookid' in str(col).lower():
            book_id_col = col
            break
    
    if book_id_col is None:
        # Создаем новую колонку book_id
        book_id_col = 'book_id'
        df[book_id_col] = ''
        print(f"  Создана новая колонка: {book_id_col}")
    else:
        print(f"  Найдена колонка book_id: {book_id_col}")
    
    # Ищем колонку с названием
    title_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'название' in col_lower or 'title' in col_lower or 'назв' in col_lower or col_lower == 'name':
            title_col = col
            break
    
    if title_col is None:
        print(f"  Предупреждение: не найдена колонка с названием на листе {sheet_name}")
        continue
    
    print(f"  Найдена колонка названия: {title_col}")
    
    # Исправляем book_id для каждой строки
    changes_count = 0
    for idx in df.index:
        author = str(df.at[idx, author_col]) if pd.notna(df.at[idx, author_col]) else ''
        title = str(df.at[idx, title_col]) if pd.notna(df.at[idx, title_col]) else ''
        
        if author and author != 'nan' and title and title != 'nan':
            # Извлекаем фамилию (первое слово из строки автора)
            author_parts = author.strip().split()
            if author_parts:
                surname = author_parts[0]  # Фамилия - первое слово
                
                # Извлекаем первое слово из названия
                title_parts = str(title).strip().split()
                if not title_parts:
                    continue
                first_word_title = title_parts[0]  # Первое слово названия
                
                # Очищаем от лишних символов (оставляем только буквы, цифры, дефисы)
                surname_clean = re.sub(r'[^\w-]', '', surname).strip()
                title_word_clean = re.sub(r'[^\w-]', '', first_word_title).strip()
                
                # Объединяем: фамилия_первое_слово_названия
                book_id_parts = [surname_clean, title_word_clean]
                book_id_parts = [p for p in book_id_parts if p]  # Убираем пустые части
                
                if book_id_parts:
                    new_book_id = '_'.join(book_id_parts)
                    # Приводим к нижнему регистру и заменяем ё на е
                    new_book_id = new_book_id.lower().replace('ё', 'е')
                    df.at[idx, book_id_col] = new_book_id
                    changes_count += 1
    
    print(f"  Обновлено строк: {changes_count}")
    
    # Сохраняем DataFrame для последующего сохранения
    dataframes[sheet_name] = df

# Сохраняем все изменения обратно в Excel файл
print("\nСохранение изменений в файл...")
with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
    for sheet_name, df in dataframes.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"  Сохранен лист: {sheet_name}")

print(f"\nФайл сохранен: {file_path}")

